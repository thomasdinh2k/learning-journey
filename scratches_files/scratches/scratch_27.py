import openpyxl,ezgmail,sys,os,smtplib

# Open the Spreadsheet and get the lastest dues status

workbook = openpyxl.load_workbook('/Users/thomas/Library/CloudStorage/OneDrive-Personal/Python Coding/Project_2_21.March/duesRecords.xlsx')
print('Loading Workbook\n')

sheet = workbook['Sheet1']
lastCol = sheet.max_column
lastestMonth = sheet.cell(row=1, column=lastCol).value
print(lastestMonth)

# Check the unpaid member status in Excel file
unpaidMember = {}
for r in range (2,sheet.max_row + 1 ):
    payStatus = sheet.cell(row=r,column=lastCol).value
    print(payStatus)
    if payStatus != 'paid':
        name = sheet.cell(row = r, column= 1).value
        email = sheet.cell(row = r , column = 2).value
        unpaidMember[name]=email #Name and Email are added to the dictonary
print('Checking unpaid member - Done!')
print(unpaidMember[name],'are unpaid!')

# Checking credentials
os.chdir("/Users/thomas/Library/CloudStorage/OneDrive-Personal/Python Coding/Credentials/Gmail")
print(f'Work directory has been changed to {os.getcwd()}')
print('Sending....')

# Sending Email in Loop
for name, email in unpaidMember.items():
    Subject = 'REMINDER: %s dues unpaid'%(lastestMonth)
    Body = 'Dear %s,\nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!'%(name,lastestMonth)
    print('Sending Email to %s.....'%(email))
    # Code to make it work
    try:
        ezgmail.send(email,Subject,Body,cc='work.dnstung@gmail.com')
        print('Email sent successfully!')
    except Exception as e:
        print(f'Error sending email: {e} \n\nPlease have a look and do not give up!')
